# SJTU EE208
INDEX_DIR = "IndexFiles.index"
import sys, os, lucene

from java.io import File
from java.nio.file import Path
from org.apache.lucene.analysis.standard import StandardAnalyzer
from org.apache.lucene.index import DirectoryReader
from org.apache.lucene.queryparser.classic import QueryParser
from org.apache.lucene.store import SimpleFSDirectory
from org.apache.lucene.search import IndexSearcher
from org.apache.lucene.util import Version
from org.apache.lucene.search import BooleanQuery
from org.apache.lucene.search import BooleanClause
from urllib.parse import urlparse
from bs4 import BeautifulSoup
import jieba
from org.apache.lucene.search.highlight import Highlighter, QueryScorer, SimpleFragmenter, SimpleHTMLFormatter

import cv2
import time
import numpy as np
import torch
import torchvision
import torchvision.transforms as transforms
from torchvision.datasets.folder import default_loader
from PIL import Image
from matplotlib import pyplot as plt

import os
import requests
import pandas as pd
import pymysql
import xlwt
from openpyxl import load_workbook

def write_url_into_xlsx(imgurl):
    workbook = load_workbook(filename='face_news/url2.xlsx')
    sheet1 = workbook["url"]
    sheet1.cell(1,1).value ="url"
    sheet1.cell(2,1).value = imgurl
    workbook.save('face_news/url2.xlsx')

def download_img():
    path="face_news/picture_compare/"   #设置图片文件路径，前提是必须要有abc这个文件夹
    df=pd.read_excel('face_news/url2.xlsx')
    urls=df['url']
    for i in range(len(urls)):
        try:
            # print(i)
            r = requests.request('get',urls[i])  #获取网页
            # print(r.status_code)
            with open(path+str("target")+str(i+1)+'.png','wb') as f:  #打开写入到path路径里-二进制文件，返回的句柄名为f
                f.write(r.content)  #往f里写入r对象的二进制文件
        except:
            pass
        # f.close()


model = torchvision.models.vgg16(pretrained=True)
def features(x,method):
    if method=='vgg16' :
        x = model.features(x)
        x = model.avgpool(x)
    else:
        x=x
    return x

def get_vgg16_feature(imgname):
    Method=['vgg16']
    for method in Method:
        # print('Load model:{}'.format(method))
        # model = torch.hub.load('pytorch/vision', method, pretrained=True)
        

        normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                        std=[0.229, 0.224, 0.225])
        trans = transforms.Compose([
            transforms.Resize(256),
            transforms.CenterCrop(224),
            transforms.ToTensor(),
            normalize,
        ])
        # print('Prepare image data!')
        #获得对比文件的feature
        figname=imgname
        # figname='apple.png'
        test_image = default_loader(figname)
        input_image = trans(test_image)
        input_image = torch.unsqueeze(input_image, 0)

        # print('Extract features!')
        start = time.time()
        image_feature = features(input_image,method)
        image_feature = image_feature.detach().numpy()
        # print('Time for extracting features: {:.2f}'.format(time.time() - start))
        # print('Save features!')
        np.save('face_news/picture_compare features/features_of_target1.npy', image_feature)


def map(m):
    if 0 <= m < 0.32:
        return 0
    elif 0.32 <= m < 0.40:
        return 1
    else:
        return 2

def vector(img, h1, h2, w1, w2):
    channels = cv2.split(img[h1:h2,w1:w2])
    #b g r
    lis = []
    sum = 0
    for i in channels:
        # print(type(i))
        lis.append(i.sum())
        sum += i.sum()
    for i in range(3):
        #lis[i]/sum 为三色素占比 b g r ，利用map分成3部分区分
        lis[i] = map(lis[i]/sum)
        #lis[i] = lis[i] / sum
    return lis  #最后得到的是一个列表，每个元素范围为0，1，2

def merge_vector(file):
    img = cv2.imread(file)
    h,w = img.shape[0],img.shape[1]
    mh = int(h/2) # medium height
    mw = int(w/2) # medium width
    res = []
    #分成四块,对应ppt中的四个区块
    res.extend(vector(img,0,mh,0,mw)) #H1
    res.extend(vector(img,0,mh, mw, w))#H2
    res.extend(vector(img, mh, h, 0, mw))#H3
    res.extend(vector(img, mh, h, mw, w))#H4
    return res

def choose_subset(dd):
    #Ii可以取1~12，我们只对有元素的Ii进行操作
    #选取子集的长度为几，得到的二进制数的长度就为几
    #以下为所取自己对应的各个Ii,此处选取的subset长度为6
    # lst=[[1],[3],[],[7,8],[],[],[],[],[],[],[],[]]
    if dd==4:
        lst=[[1],[3],[],[7,8],[],[],[],[],[],[],[],[]]
    else:
        lst=[[1],[3],[],[7,8],[],[],[14],[],[18],[],[],[]]
    return lst

def LSHsearch(vector,dd):#需要返回一个bin的大小
    tmp=[0]*24    #对应的是xi,i from 1 to 24
    #d=12,c=2,d'=24
    d=12
    c=2
    for i in range(len(vector)):
        if vector[i]==0:
            tmp[2*i]=0
            tmp[2*i+1]=0
        elif vector[i]==1:
            tmp[2*i]=1
            tmp[2*i+1]=0
        else:
            tmp[2*i]=1
            tmp[2*i+1]=1

    lst=choose_subset(dd)
    hash_res_lst=[]
    # compare_lst=[x for x in lst if x!=[]]
    for i in range(d):
        if lst[i]==[]:
            continue
        else:
            for item in lst[i]:
                if (item-c*i)<=vector[i]:
                    hash_res_lst.append(1)
                else:
                    hash_res_lst.append(0)
    hashres=0
    # print(hash_res_lst)
    for i in range(len(hash_res_lst)):
        hashres+=hash_res_lst[i]*(2**(dd-1-i))
    return hashres

def cal_vector_angle(A,B):#计算向量之间的夹角
    x=np.linalg.norm(A)
    y=np.linalg.norm(B)
    z=float(np.sum(np.multiply(A,B))/(x*y))
    return z

def get_array_of_fig(tmp_load):
    len1=tmp_load.shape[1]
    len2=tmp_load.shape[2]
    len3=tmp_load.shape[3]
    len=len1*len2*len3
    lst=[]
    for i in range(len1):
        for j in range(len2):
            for k in range(len3):
                # lst[i]=tmp_load[0][i][j][k]
                lst.append(tmp_load[0][i][j][k])
    array=np.array(lst)
    return array

def get_compare():   #得到对比图片的npy文件
    compare_fig=np.load('face_news/picture_compare features/features_of_target1.npy')
    compare_array=get_array_of_fig(compare_fig)
    return compare_array

def vector_method(aimset):  #aimset是得到的对应的hash的列表
    loadData=[]#loadData用于存储test图片中对应的归一化后的向量，每种模型对应的向量长度一致
    for i in aimset:
        try:
            tmp_load=np.load('face_news/picture features/feature_of_img{0}.npy'.format(i))
            tmp_array=get_array_of_fig(tmp_load)###
            loadData.append(tmp_array)
        except:
            pass

    compare_array=get_compare()
    compare={}#用于储存最后得到的匹配信息
    compare_lst=[]

    for i in range(len(loadData)):
        value=cal_vector_angle(loadData[i], compare_array)
        compare[value]=aimset[i]
        compare_lst.append(value)
    compare_lst.sort(reverse=True)#求得的是cos的值，越大越相似
   
    resnum=[]
    for i in range(min(10,len(loadData))):
        resnum.append(compare[compare_lst[i]])
    return resnum,compare_lst


def LSH_search(imgname):
    # t1 = time.time()
    target=imgname
    vec_target=merge_vector(target)
    hashres_target = LSHsearch(vec_target,4)   ##这个是计算输入图片得到的hash值
    # print(hashres_target)
    # x=PreProcessLSH('Dataset')
    # print(x)
    aimset_load=np.load("face_news/hash.npy",allow_pickle =True).item() #将装有hash的npy文件导入，得到的是一个列表套列表
    # print(aimset_load)
    aimset0=aimset_load[int(hashres_target)]
    # print(aimset0)   #zhege is true

    res0,res_value0=vector_method(aimset0)
    return res0

def findimgInfo(Nos):#Nos是检索结果的列表，里面的元素是每一个结果图片的编号numbers-1
    loader=np.load("process.npy",allow_pickle=True).item()
    results=[]
    for i in Nos:
        number=i+1
        info=dict()
        inf_lst=loader[number]
        info["news_id"]=inf_lst[0]
        info["imgurl"]=inf_lst[1]
        info["title"]=inf_lst[2]
        info["url"]=inf_lst[3]
        info["date"]=inf_lst[4]
        results.append(info)
    return results


# image_url1="https://p3.itc.cn/q_70/images03/20220112/be81e28f8dc640319c29e86d45496a86.jpeg"
# image_url1="https://p9.itc.cn/q_70/images03/20220114/6062a50b69fc45ecb7825a988f95534a.png"
def picture_to_picture(image_url):
    write_url_into_xlsx(image_url)
    download_img()
    imgname="face_news/picture_compare/target1.png"
    get_vgg16_feature(imgname)
    tmp_list=LSH_search(imgname)
    result=findimgInfo(tmp_list)
    # print(tmp_list)
    return result

# ppp=picture_to_picture(image_url1)
# print(ppp)