import pandas as pd
import pymysql
import xlwt
from openpyxl import load_workbook

workbook = load_workbook(filename='face_news/url1.xlsx')
sheet1 = workbook["url"]
sheet1.cell(1,1).value ="url"

# 创建数据库的连接 #
db = pymysql.connect(host='152.136.97.17',
                     port=3306,
                     user='cjx',
                     password='111111',
                     db='crawler_pages',
                             )

cursor = db.cursor()

sql = "SELECT * FROM imgInfo LIMIT 100;" 
cursor.execute(sql)
data = cursor.fetchall()

for i in range(len(data)):
    sheet1.cell(i+2,1).value=data[i][1]


workbook.save('face_news/url1.xlsx')